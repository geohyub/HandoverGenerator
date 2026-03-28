"""Excel readiness report generation."""

from pathlib import Path
from typing import Optional, Tuple

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from handover_check.i18n import normalize_lang, severity_text, status_text, tr
from handover_check.insights import build_readiness_insight
from handover_check.models import FolderResult, ResultStatus, ValidationReport
from handover_check.validators.total_size import format_size


TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
SECTION_TITLE_FONT = Font(bold=True, color="1F1F1F", size=12)
SUBTITLE_FONT = Font(bold=True, color="1F1F1F", size=11)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
TITLE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SKIP_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
INFO_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
BLOCKER_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
MAJOR_FILL = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
MINOR_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

STATUS_FILLS = {
    ResultStatus.PASS: PASS_FILL,
    ResultStatus.FAIL: FAIL_FILL,
    ResultStatus.WARNING: WARN_FILL,
    ResultStatus.SKIP: SKIP_FILL,
    ResultStatus.INFO: INFO_FILL,
}

SEVERITY_FILLS = {
    "Blocker": BLOCKER_FILL,
    "Major": MAJOR_FILL,
    "Minor": MINOR_FILL,
    "Info": INFO_FILL,
}


def _report_lang(report: ValidationReport) -> str:
    return normalize_lang(getattr(report, "language", "en"))


def _write_cell(ws, row, column, value, fill=None, font=None, alignment=None):
    cell = ws.cell(row=row, column=column, value=value)
    cell.border = THIN_BORDER
    cell.font = font or BODY_FONT
    cell.alignment = alignment or WRAP_ALIGNMENT
    if fill:
        cell.fill = fill
    return cell


def _write_header(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        _write_cell(
            ws,
            row,
            col,
            header,
            fill=HEADER_FILL,
            font=HEADER_FONT,
            alignment=Alignment(horizontal="center", vertical="center"),
        )


def _section_title(ws, row, title, end_col=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_TITLE_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER


def _auto_width(ws, min_width=10, max_width=60):
    for col in ws.iter_cols():
        first_real_cell = next(
            (cell for cell in col if not isinstance(cell, MergedCell)),
            None,
        )
        if first_real_cell is None:
            continue
        max_len = 0
        col_letter = get_column_letter(first_real_cell.column)
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _extract_count_info(folder_result: FolderResult) -> Tuple[str, str]:
    for rule_result in folder_result.rule_results:
        if rule_result.rule_type != "count_match":
            continue
        message = rule_result.message
        if "/" not in message:
            continue
        left, right = message.split("/", 1)
        found = left.strip()
        expected = right.split(" ", 1)[0].strip()
        return found, expected
    return "", ""


def _sheet_title(ws, title):
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 24


def _write_summary_tab(wb, report: ValidationReport):
    lang = _report_lang(report)
    insight = build_readiness_insight(report, lang)
    section_lookup = {
        section["scope"]: section for section in insight.section_overview
    }
    ws = wb.active
    ws.title = tr("sheet_summary", lang)
    _sheet_title(ws, tr("summary_title", lang))

    row = 3
    _section_title(ws, row, tr("section_executive_readiness", lang))
    row += 1

    summary_rows = [
        (tr("readiness_label", lang), insight.readiness_label),
        (tr("readiness_score", lang), f"{insight.readiness_percent}%"),
        (tr("overall_status", lang), status_text(report.overall_status.value, lang)),
        (tr("headline", lang), insight.headline),
        (tr("project", lang), report.project or tr("na", lang)),
        (tr("client", lang), report.client or tr("na", lang)),
        (tr("profile", lang), report.profile_name),
        (tr("delivery_path", lang), report.delivery_path),
        (tr("validated_on", lang), report.timestamp),
        (tr("total_files", lang), f"{report.total_files:,}"),
        (tr("total_size", lang), format_size(report.total_size_bytes)),
        (
            tr("check_mix", lang),
            f"{insight.passed_checks} {tr('passed', lang)}, "
            f"{insight.warning_checks} {tr('warnings', lang)}, "
            f"{insight.failed_checks} {tr('failed', lang)}, "
            f"{insight.skipped_checks} {tr('skipped', lang)}",
        ),
        (
            tr("severity_mix", lang),
            f"{insight.blocker_count} {severity_text('Blocker', lang)}, "
            f"{insight.major_count} {severity_text('Major', lang)}, "
            f"{insight.minor_count} {severity_text('Minor', lang)}",
        ),
    ]
    for label, value in summary_rows:
        _write_cell(ws, row, 1, label, fill=SECTION_FILL, font=SUBTITLE_FONT)
        _write_cell(ws, row, 2, value)
        row += 1

    if insight.coverage:
        coverage = insight.coverage
        _write_cell(ws, row, 1, tr("line_coverage", lang), fill=SECTION_FILL, font=SUBTITLE_FONT)
        _write_cell(
            ws,
            row,
            2,
            f"{coverage.fully_covered_lines}/{coverage.expected_lines} "
            f"{tr('fully_covered_lines', lang).lower()} ({coverage.fully_covered_percent}%)",
        )
        row += 1

    row += 1
    _section_title(ws, row, tr("section_validation_basis", lang))
    row += 1
    basis_rows = [
        (tr("base_profile", lang), report.base_profile or tr("direct_profile", lang)),
        (tr("profile_description", lang), report.profile_description or tr("na", lang)),
        (tr("profile_notes", lang), report.profile_notes or tr("na", lang)),
        (tr("line_list_source", lang), report.line_list_source or tr("not_used", lang)),
        (tr("line_id_column", lang), report.line_id_column or tr("na", lang)),
        (
            tr("line_filter", lang),
            (
                f"{report.line_status_column} == {report.line_status_filter}"
                if report.line_status_column and report.line_status_filter
                else tr("no_status_filter", lang)
            ),
        ),
        (
            tr("resolved_variables", lang),
            ", ".join(
                f"{key}={value or '<empty>'}"
                for key, value in sorted(report.resolved_variables.items())
            ) or tr("none", lang),
        ),
    ]
    for label, value in basis_rows:
        _write_cell(ws, row, 1, label, fill=SECTION_FILL, font=SUBTITLE_FONT)
        _write_cell(ws, row, 2, value)
        row += 1

    row += 1
    _section_title(ws, row, tr("section_next_actions", lang))
    row += 1
    if insight.top_actions:
        for action in insight.top_actions:
            _write_cell(ws, row, 1, tr("action_label", lang), fill=SECTION_FILL, font=SUBTITLE_FONT)
            _write_cell(ws, row, 2, action)
            row += 1
    else:
        _write_cell(ws, row, 1, tr("action_label", lang), fill=SECTION_FILL, font=SUBTITLE_FONT)
        _write_cell(ws, row, 2, tr("no_corrective_actions", lang))
        row += 1

    row += 1
    _section_title(ws, row, tr("section_trust_notes", lang))
    row += 1
    if insight.trust_notes:
        for note in insight.trust_notes:
            _write_cell(ws, row, 1, tr("note_label", lang), fill=SECTION_FILL, font=SUBTITLE_FONT)
            _write_cell(ws, row, 2, note)
            row += 1
    else:
        _write_cell(ws, row, 1, tr("note_label", lang), fill=SECTION_FILL, font=SUBTITLE_FONT)
        _write_cell(ws, row, 2, tr("no_trust_caveats", lang))
        row += 1

    row += 1
    _section_title(ws, row, tr("section_overview", lang))
    row += 1
    headers = [
        tr("col_section", lang),
        tr("col_description", lang),
        tr("col_required", lang),
        tr("col_status", lang),
        tr("col_files_found", lang),
        tr("col_files_expected", lang),
        tr("col_primary_finding", lang),
    ]
    _write_header(ws, headers, row=row)
    row += 1

    for folder_result in report.folder_results:
        files_found, files_expected = _extract_count_info(folder_result)
        if not files_found:
            file_count = sum(
                1
                for file_info in report.file_inventory
                if file_info.folder == folder_result.path
                or file_info.folder.startswith(folder_result.path + "/")
            )
            files_found = str(file_count) if file_count else ""
        overview = section_lookup.get(folder_result.path, {})
        primary_finding = overview.get("primary_message") or tr("no_issues_raised", lang)

        _write_cell(ws, row, 1, folder_result.path)
        _write_cell(ws, row, 2, folder_result.description or tr("na", lang))
        _write_cell(ws, row, 3, tr("no", lang) if folder_result.optional else tr("yes", lang))
        _write_cell(
            ws,
            row,
            4,
            status_text(folder_result.status.value, lang),
            fill=STATUS_FILLS.get(folder_result.status, INFO_FILL),
        )
        _write_cell(ws, row, 5, files_found)
        _write_cell(ws, row, 6, files_expected)
        _write_cell(ws, row, 7, primary_finding)
        row += 1

    _auto_width(ws)


def _write_issues_tab(wb, report: ValidationReport):
    lang = _report_lang(report)
    insight = build_readiness_insight(report, lang)
    ws = wb.create_sheet(tr("sheet_issues", lang))
    _sheet_title(ws, tr("issues_title", lang))

    row = 3
    _section_title(ws, row, tr("section_issue_summary", lang), end_col=5)
    row += 1
    summary_rows = [
        (tr("total_issues", lang), len(insight.issues)),
        (severity_text("Blocker", lang), insight.blocker_count),
        (severity_text("Major", lang), insight.major_count),
        (severity_text("Minor", lang), insight.minor_count),
        (tr("impacted_sections", lang), insight.impacted_sections),
    ]
    for label, value in summary_rows:
        _write_cell(ws, row, 1, label, fill=SECTION_FILL, font=SUBTITLE_FONT)
        _write_cell(ws, row, 2, value)
        row += 1

    row += 1
    headers = [
        "#",
        tr("col_severity", lang),
        tr("col_section", lang),
        tr("col_check", lang),
        tr("col_category", lang),
        tr("col_status", lang),
        tr("col_finding", lang),
        tr("recommendation", lang),
        tr("col_evidence", lang),
    ]
    _write_header(ws, headers, row=row)
    row += 1

    if insight.issues:
        for index, issue in enumerate(insight.issues, 1):
            evidence_text = "\n".join(issue.details[:12])
            if len(issue.details) > 12:
                evidence_text += f"\n... +{len(issue.details) - 12} more"

            _write_cell(ws, row, 1, index)
            _write_cell(
                ws,
                row,
                2,
                severity_text(issue.severity, lang),
                fill=SEVERITY_FILLS.get(issue.severity, INFO_FILL),
            )
            _write_cell(ws, row, 3, issue.scope)
            _write_cell(ws, row, 4, issue.check_label)
            _write_cell(ws, row, 5, issue.category)
            _write_cell(
                ws,
                row,
                6,
                status_text(issue.status.value, lang),
                fill=STATUS_FILLS.get(issue.status, INFO_FILL),
            )
            _write_cell(ws, row, 7, issue.message)
            _write_cell(ws, row, 8, issue.action_hint)
            _write_cell(ws, row, 9, evidence_text or tr("see_message", lang))
            row += 1
    else:
        _write_cell(ws, row, 1, tr("no_issues_found", lang))

    _auto_width(ws)


def _write_line_coverage_tab(wb, report: ValidationReport):
    lang = _report_lang(report)
    ws = wb.create_sheet(tr("sheet_line_coverage", lang))
    _sheet_title(ws, tr("line_coverage_title", lang))

    insight = build_readiness_insight(report, lang)
    coverage = insight.coverage
    row = 3

    if not report.line_coverage or not coverage:
        _write_cell(
            ws,
            row,
            1,
            tr("no_line_coverage", lang),
            fill=SECTION_FILL,
            font=SUBTITLE_FONT,
        )
        return

    _section_title(ws, row, tr("section_coverage_summary", lang))
    row += 1
    summary_rows = [
        (tr("expected_lines", lang), coverage.expected_lines),
        (tr("tracked_folders", lang), coverage.tracked_folders),
        (
            tr("fully_covered_lines", lang),
            f"{coverage.fully_covered_lines}/{coverage.expected_lines} ({coverage.fully_covered_percent}%)",
        ),
        (
            tr("missing_lines", lang),
            ", ".join(coverage.missing_lines[:10]) or tr("none", lang),
        ),
    ]
    for label, value in summary_rows:
        _write_cell(ws, row, 1, label, fill=SECTION_FILL, font=SUBTITLE_FONT)
        _write_cell(ws, row, 2, value)
        row += 1

    row += 1
    _section_title(ws, row, tr("section_matrix", lang), end_col=1 + len(report.line_coverage["folders"]))
    row += 1

    folders = report.line_coverage["folders"]
    lines = report.line_coverage["lines"]
    matrix = report.line_coverage["matrix"]

    headers = [tr("col_line", lang)] + folders
    _write_header(ws, headers, row=row)
    row += 1

    for line in lines:
        _write_cell(ws, row, 1, line)
        for column_index, folder in enumerate(folders, 2):
            found = matrix.get(line, {}).get(folder, False)
            _write_cell(
                ws,
                row,
                column_index,
                tr("ok", lang) if found else tr("missing", lang),
                fill=PASS_FILL if found else FAIL_FILL,
                alignment=Alignment(horizontal="center", vertical="center"),
            )
        row += 1

    _auto_width(ws)


def _write_file_inventory_tab(wb, report: ValidationReport):
    lang = _report_lang(report)
    ws = wb.create_sheet(tr("sheet_file_inventory", lang))
    _sheet_title(ws, tr("file_inventory_title", lang))
    row = 3
    headers = [
        "#",
        tr("col_folder", lang),
        tr("col_filename", lang),
        tr("col_size_mb", lang),
        tr("col_modified", lang),
        tr("col_naming_ok", lang),
        tr("col_notes", lang),
    ]
    _write_header(ws, headers, row=row)
    row += 1

    for index, file_info in enumerate(report.file_inventory, 1):
        _write_cell(ws, row, 1, index)
        _write_cell(ws, row, 2, file_info.folder)
        _write_cell(ws, row, 3, file_info.filename)
        _write_cell(ws, row, 4, round(file_info.size_bytes / (1024 * 1024), 2))
        _write_cell(ws, row, 5, file_info.modified)

        naming_value = tr("na", lang)
        naming_fill = INFO_FILL
        if file_info.naming_ok is True:
            naming_value = tr("ok", lang)
            naming_fill = PASS_FILL
        elif file_info.naming_ok is False:
            naming_value = tr("no", lang)
            naming_fill = FAIL_FILL
        _write_cell(
            ws,
            row,
            6,
            naming_value,
            fill=naming_fill,
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        _write_cell(ws, row, 7, file_info.notes)
        row += 1

        if index >= 10000:
            _write_cell(ws, row, 1, "...")
            _write_cell(
                ws,
                row,
                2,
                tr("truncated_inventory", lang, count=len(report.file_inventory)),
            )
            break

    _auto_width(ws)


def _write_naming_violations_tab(wb, report: ValidationReport):
    lang = _report_lang(report)
    ws = wb.create_sheet(tr("sheet_naming_violations", lang))
    _sheet_title(ws, tr("naming_violations_title", lang))
    row = 3
    headers = [
        "#",
        tr("col_folder", lang),
        tr("col_filename", lang),
        tr("col_expected_pattern", lang),
        tr("col_issue", lang),
    ]
    _write_header(ws, headers, row=row)
    row += 1

    if report.naming_violations:
        for index, violation in enumerate(report.naming_violations, 1):
            _write_cell(ws, row, 1, index)
            _write_cell(ws, row, 2, violation.get("folder", ""))
            _write_cell(ws, row, 3, violation.get("filename", ""))
            _write_cell(ws, row, 4, violation.get("expected_pattern", ""))
            _write_cell(ws, row, 5, violation.get("issue", ""))
            row += 1
    else:
        _write_cell(ws, row, 1, tr("no_issues_found", lang))

    _auto_width(ws)


def generate_excel_report(report: ValidationReport, output_path: Path) -> None:
    """Generate the full Excel report."""

    wb = Workbook()
    _write_summary_tab(wb, report)
    _write_issues_tab(wb, report)
    _write_line_coverage_tab(wb, report)
    _write_file_inventory_tab(wb, report)
    _write_naming_violations_tab(wb, report)
    wb.save(str(output_path))
