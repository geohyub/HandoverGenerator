"""Tests for report writers."""

import shutil

from openpyxl import load_workbook

from handover_check.engine import ValidationEngine
from handover_check.reporters.excel import generate_excel_report


def test_generate_excel_report_contains_readiness_sections(
    tmp_delivery,
    tmp_linelist,
    test_profile_yaml,
    tmp_path,
):
    shutil.copy(tmp_linelist, tmp_delivery / "line_list.csv")

    engine = ValidationEngine(
        delivery_path=tmp_delivery,
        profile_path=test_profile_yaml,
        cli_vars={"project_code": "TEST2025", "area_code": "RS"},
    )
    report = engine.run()

    output_path = tmp_path / "report.xlsx"
    generate_excel_report(report, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "Summary",
        "Issues",
        "Line Coverage",
        "File Inventory",
        "Naming Violations",
    ]

    summary = workbook["Summary"]
    issues = workbook["Issues"]

    assert summary["A1"].value == "Handover Package Readiness Summary"
    assert summary["A3"].value == "Executive Readiness"
    assert summary["A4"].value == "Readiness label"
    assert issues["A1"].value == "Corrective Action Register"
    assert issues["A3"].value == "Issue Summary"


def test_generate_excel_report_uses_korean_when_requested(
    tmp_delivery,
    tmp_linelist,
    test_profile_yaml,
    tmp_path,
):
    shutil.copy(tmp_linelist, tmp_delivery / "line_list.csv")

    engine = ValidationEngine(
        delivery_path=tmp_delivery,
        profile_path=test_profile_yaml,
        cli_vars={"project_code": "TEST2025", "area_code": "RS"},
        language="ko",
    )
    report = engine.run()

    output_path = tmp_path / "report_ko.xlsx"
    generate_excel_report(report, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "요약",
        "이슈",
        "라인 커버리지",
        "파일 인벤토리",
        "이름 위반",
    ]

    summary = workbook["요약"]
    issues = workbook["이슈"]

    assert summary["A1"].value == "납품 준비도 요약"
    assert summary["A3"].value == "핵심 준비도"
    assert issues["A1"].value == "시정 조치 등록부"
